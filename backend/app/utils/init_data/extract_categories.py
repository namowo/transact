"""Regenerates app/utils/init_data/data/*.json from the source xlsx files.

Run with `uv run python app/utils/init_data/extract_categories.py` (from
backend/). Only needs re-running if Categories.xlsx or the Definitionen
file change - the JSON output is what categories.py actually loads at
app startup, not these spreadsheets.

Each sheet/column is mapped by hand onto the DB model it actually
belongs to, since several Excel names differ from their model:
- 05bb_ItemPartMaterialCategori -> SurfaceMaterialCategory (not a
  "condition" table despite living next to ItemPartsCategory)
- Items/Condition rows in the Definitionen "Enums with Descriptions"
  sheet -> ConditionOfItemPartCategory
- Individuals/DNAsheddingPropensity enum rows -> DNASheddingPropensityCategory
- 07aa_TypeOfSwabCategories.SupplierCategoryID is a 1-based row index
  into 07aaa_SupplierCategories, resolved here to a supplier name and
  resolved again to an id at import time in categories.py
- 05ba_ItemPartsCategories.ItemCategoryID / 06bab_ItemSubcategories.
  ItemCategoryID are resolved the same way, to an item_category name
- 06ab_DeterminationOfSheddingPro packs several non-atomic fields
  (Authors, RestrictionsPriorToSampling, MonitoredTransferFactors,
  ShedderTest) into single free-text cells. These are split on commas
  into separate lookup rows; a trailing "(N min)"/"(N s)" is parsed as
  a duration attached to the row as a whole (it's ambiguous which
  comma-separated phrase it belongs to when there's more than one -
  those rows are printed as warnings for manual review after import)
- Cut-OutMethods/CuttingDevice, Results/Degradation, Results/Inhibition,
  EPGInterpretationMethods/ApplicationAnalyticalThreshold and
  EPGInterpretationMethods/StutterFilter enum rows (Definitionen file)
  -> cutting_device, degradation_category, inhibition_category,
  application_analytical_threshold, stutter_filter lookup tables
"""

import json
import re
from pathlib import Path

import openpyxl

INIT_DATA_DIR = Path(__file__).resolve().parent
OUT_DIR = INIT_DATA_DIR / "data"
OUT_DIR.mkdir(exist_ok=True)

CATEGORIES_XLSX = INIT_DATA_DIR / "Categories.xlsx"
DEFINITIONS_XLSX = INIT_DATA_DIR / "Definitionen_Auswahlmöglichkeiten_Repositorx.xlsx"


def clean(v):
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def rows(ws):
    """Yield non-empty rows as dicts keyed by the header row, values cleaned."""
    it = ws.iter_rows(values_only=True)
    header = [clean(h) for h in next(it)]
    for r in it:
        r = [clean(v) for v in r]
        if all(v is None for v in r):
            continue
        yield dict(zip(header, r))


def dump(name, records):
    path = OUT_DIR / f"{name}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"wrote {len(records)} records -> {path}")


def extract():
    wb = openpyxl.load_workbook(CATEGORIES_XLSX, data_only=True)

    # --- simple name-only / name+description sheets -> flat lookup tables ---
    simple_sheet_to_table = {
        "02a_ScenarioCategories": "scenario_category",
        "02d_DisturbanceCategories": "disturbance_category",
        "03a_ActivityCategories": "activity_category",
        "04s_SourceOfDNACategories": "source_of_dna_category",
        "05aa_LocationOfBodyCategories": "location_of_body_category",
        "05ab_BodyPartConditionCategorie": "body_part_condition_category",
        "06ba_ItemCategories": "item_category",
        "07aaa_SupplierCategories": "supplier",  # single-column subset, see below
    }
    for sheet, table in simple_sheet_to_table.items():
        ws = wb[sheet]
        records = []
        for row in rows(ws):
            header_key = next(iter(row))
            name = row[header_key]
            if name is None:
                continue
            rec = {"name": name}
            if "Description" in row and row["Description"] is not None:
                rec["description"] = row["Description"]
            records.append(rec)
        dump(table, records)

    # --- sheets with an explicit Description column already handled above ---
    for sheet, table in {
        "02e_GeographicLocationCategorie": "geographic_location_category",
        "07ab_SwabbingTechniqueCategorie": "swabbing_technique_category",
        "08a_PrincipleOfExtractionMethod": "principle_of_extraction_method_category",
    }.items():
        ws = wb[sheet]
        records = []
        for row in rows(ws):
            header_key = next(iter(row))
            name = row.get(header_key)
            if name is None:
                continue
            records.append({"name": name, "description": row.get("Description")})
        dump(table, records)

    # --- 05bb_ItemPartMaterialCategori -> surface_material_category ---
    ws = wb["05bb_ItemPartMaterialCategori"]
    records = [{"name": row["ItemPartMaterialCategory"]} for row in rows(ws)]
    dump("surface_material_category", records)

    # --- 06ba_ItemCategories, needed to resolve item_category_id below ---
    item_category_names = [
        row["ItemCategory"] for row in rows(wb["06ba_ItemCategories"])
    ]

    # --- 05ba_ItemPartsCategories -> item_parts_category (name, item_category_name) ---
    ws = wb["05ba_ItemPartsCategories"]
    seen = {}
    for row in rows(ws):
        name = row["ItemPartsCategory"]
        item_category_name = row.get("ItemCategoryID")
        if name is not None and name not in seen:
            seen[name] = item_category_name
    dump(
        "item_parts_category",
        [
            {"name": name, "item_category_name": item_category_name}
            for name, item_category_name in seen.items()
        ],
    )

    # --- 06bab_ItemSubcategories -> item_subcategory (name, item_category_name) ---
    # ItemCategoryID here is a 1-based row index into 06ba_ItemCategories.
    ws = wb["06bab_ItemSubcategories"]
    records = []
    for row in rows(ws):
        name = row.get("ItemSubcategory")
        if name is None:
            continue
        category_id = row.get("ItemCategoryID")
        item_category_name = (
            item_category_names[int(category_id) - 1]
            if category_id and 0 < int(category_id) <= len(item_category_names)
            else None
        )
        records.append({"name": name, "item_category_name": item_category_name})
    dump("item_subcategory", records)

    # --- 06aa_SkinDiseaseCategories -> skin_disease_category ---
    ws = wb["06aa_SkinDiseaseCategories"]
    bool_map = {"increase": True, "decrease": False, "none": False}
    records = []
    for row in rows(ws):
        name = row["SkinDiseaseCategory"]
        if name is None:
            continue
        influence_raw = (row.get("InfluenceOnSheddingPropensity") or "").strip().lower() if row.get("InfluenceOnSheddingPropensity") else None
        records.append(
            {
                "name": name,
                "influence_on_shedding_propensity": bool_map.get(influence_raw),
                "literature": row.get("Literature"),
            }
        )
    dump("skin_disease_category", records)

    # --- 06ab_DeterminationOfSheddingPro -> determination_of_shedding_propensity_category ---
    # Authors/RestrictionsPriorToSampling/MonitoredTransferFactors/ShedderTest
    # are free text, not atomic - see module docstring. Split into rows for
    # the new lookup tables here; categories.py resolves the *_name fields.
    DURATION_RE = re.compile(r"\(\s*([\d.]+)\s*(min|s|sec|seconds?|minutes?)\s*\)")

    def parse_duration(text):
        match = DURATION_RE.search(text)
        if not match:
            return None, text
        value, unit = match.groups()
        value = float(value)
        seconds = value * 60 if unit.startswith("min") else value
        remainder = (text[: match.start()] + text[match.end() :]).strip(" ,")
        return f"PT{seconds:g}S", remainder

    def split_phrases(text):
        """Best-effort split of a free-text cell into atomic phrases."""
        duration, remainder = parse_duration(text)
        phrases = [p.strip() for p in remainder.split(",") if p.strip()]
        if len(phrases) > 1 and duration is not None:
            print(
                f"WARNING: ambiguous duration {duration!r} in {text!r} - "
                f"attached to last phrase {phrases[-1]!r}, please review"
            )
        return phrases, duration

    def parse_author_names(text):
        # "Last, First; Last2, First2; ..." - some rows repeat the same
        # author as both "Last, F." and "Last, First"; dedupe on
        # (last_name, first initial), keeping the fuller spelling.
        by_key = {}
        for part in text.split(";"):
            part = part.strip()
            if not part or "," not in part:
                continue
            last, _, first = part.partition(",")
            last, first = last.strip(), first.strip()
            if not last or not first:
                continue
            key = (last.lower(), first[0].lower())
            existing = by_key.get(key)
            if existing is None or len(first) > len(existing["first_name"]):
                by_key[key] = {"first_name": first, "last_name": last}
        return list(by_key.values())

    ws = wb["06ab_DeterminationOfSheddingPro"]
    records = []
    for row in rows(ws):
        authors_raw = row.get("Authors")
        if authors_raw is None or authors_raw == "not determined":
            continue

        restrictions, restriction_duration = split_phrases(
            row.get("RestrictionsPriorToSampling") or ""
        )
        transfer_factors, _ = split_phrases(row.get("MonitoredTransferFactors") or "")
        shedder_tests, shedder_test_duration = split_phrases(
            row.get("ShedderTest") or ""
        )

        number_of_participants = row.get("NumberOfParticipants")
        replicates = row.get("Replicates")
        records.append(
            {
                "title": row.get("Title"),
                "doi": row.get("doi"),
                "authors": parse_author_names(authors_raw),
                "restrictions": [
                    {"name": name, "duration": restriction_duration}
                    for name in restrictions
                ],
                "monitored_transfer_factor_names": transfer_factors,
                "number_of_participants": (
                    int(number_of_participants)
                    if number_of_participants
                    and number_of_participants.isdigit()
                    else None
                ),
                "replicates": (
                    int(replicates) if replicates and replicates.isdigit() else None
                ),
                "shedder_tests": [
                    {"name": name, "duration": shedder_test_duration}
                    for name in shedder_tests
                ],
                "classification_criteria_name": row.get("ClassificationCriteria"),
                "classification_scheme_name": row.get("ClassificationScheme"),
                "classification_outcome": row.get("ClassificationOutcome"),
            }
        )
    dump("determination_of_shedding_propensity_category", records)

    # --- 09a_PrincipleOfQuantMethodCateg -> principle_of_quant_method_category ---
    ws = wb["09a_PrincipleOfQuantMethodCateg"]
    records = [
        {"name": row["PrincipleOfQuantificationMethodCategory"]} for row in rows(ws)
    ]
    dump("principle_of_quant_method_category", records)

    # --- 07aaa_SupplierCategories -> supplier (name only) ---
    ws = wb["07aaa_SupplierCategories"]
    supplier_names = [row["SupplierCategory"] for row in rows(ws)]
    dump("supplier", [{"name": n} for n in supplier_names])

    # --- 07aa_TypeOfSwabCategories -> type_of_swab_category ---
    ws = wb["07aa_TypeOfSwabCategories"]
    records = []
    for row in rows(ws):
        name = row.get("TypeOfSwabCategory")
        if name is None:
            continue
        supplier_id = row.get("SupplierCategoryID")
        supplier_name = (
            supplier_names[int(supplier_id) - 1]
            if supplier_id and 0 < int(supplier_id) <= len(supplier_names)
            else None
        )
        records.append(
            {
                "name": name,
                "description": row.get("Description"),
                "catalogue_number_of_supplier": row.get("CatalogueNumberOfSupplier"),
                "full_name_as_by_supplier": row.get("FullNameAsBySupplier"),
                "supplier_name": supplier_name,  # resolved to supplier_id at import time
            }
        )
    dump("type_of_swab_category", records)

    # --- Definitionen file: Enums with Descriptions -> a few real lookup tables ---
    wb2 = openpyxl.load_workbook(DEFINITIONS_XLSX, data_only=True)
    ws = wb2["Enums with Descriptions"]
    enum_rows = list(rows(ws))

    def enum_group(entity, name):
        return [
            r["enum-options"]
            for r in enum_rows
            if r.get("Entity") == entity and r.get("Name") == name and r.get("enum-options")
        ]

    dump("sex", [{"name": v} for v in enum_group("Individuals", "Sex")])
    dump(
        "experience_level",
        [{"name": v} for v in enum_group("Recoveries", "ExperienceLevelOfSampler")],
    )
    dump(
        "dna_shedding_propensity_category",
        [{"name": v} for v in enum_group("Individuals", "DNAsheddingPropensity")],
    )
    condition_records = [
        {"name": r["enum-options"], "description": r.get("Description")}
        for r in enum_rows
        if r.get("Entity") == "Items" and r.get("Name") == "Condition"
    ]
    dump("condition_of_item_part_category", condition_records)

    dump(
        "cutting_device",
        [{"name": v} for v in enum_group("Cut-OutMethods", "CuttingDevice")],
    )
    dump(
        "degradation_category",
        [{"name": v} for v in enum_group("Results", "Degradation")],
    )
    dump(
        "inhibition_category",
        [{"name": v} for v in enum_group("Results", "Inhibition")],
    )
    dump(
        "application_analytical_threshold",
        [
            {"name": v}
            for v in enum_group(
                "EPGInterpretationMethods", "ApplicationAnalyticalThreshold"
            )
        ],
    )
    dump(
        "stutter_filter",
        [{"name": v} for v in enum_group("EPGInterpretationMethods", "StutterFilter")],
    )


if __name__ == "__main__":
    extract()
